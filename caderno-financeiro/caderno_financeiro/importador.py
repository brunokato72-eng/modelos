"""Import de planilha antiga (CSV/XLSX).

Detecta as colunas pelo nome (ignorando acento, caixa e separador), normaliza
categorias/tipos/formas pra grafia canônica, aplica as regras determinísticas no
que ficou em branco e descarta duplicatas (mesmo valor + data a até 1 dia de
distância, como definido no briefing).

Cada linha da planilha vira um lançamento — o import não re-parcela nada: se a
planilha já tem uma linha por parcela, é isso que entra.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from . import config, db, regras
from .datas import dias_entre, validar_iso
from .texto import normalizar_chave
from .valores import para_centavos, parsear_valor

# Sinônimos de cabeçalho, do mais específico pro mais genérico.
SINONIMOS: Sequence[Tuple[str, Sequence[str]]] = (
    ("valorTotal", ("valor total", "total da compra", "valor da compra", "valor bruto")),
    ("totalParcelas", ("total parcelas", "qtd parcelas", "quantidade de parcelas",
                       "numero de parcelas", "num parcelas", "parcelas", "parcela")),
    ("formaPagamento", ("forma pagamento", "forma de pagamento", "meio de pagamento",
                        "metodo de pagamento", "metodo pagamento", "pagamento",
                        "payment method", "forma")),
    ("data", ("data", "data do lancamento", "data lancamento", "data da compra",
              "data compra", "data pagamento", "dia", "date", "competencia", "vencimento")),
    ("tipo", ("tipo", "tipo lancamento", "tipo de lancamento", "natureza",
              "entrada saida", "entrada ou saida", "type")),
    ("categoria", ("categoria", "categorias", "category", "classificacao", "grupo")),
    ("conta", ("conta", "cartao", "cartao conta", "conta cartao", "banco", "carteira",
               "origem", "account", "instituicao")),
    ("descricao", ("descricao", "description", "detalhe", "detalhes", "historico",
                   "estabelecimento", "titulo", "observacao", "observacoes", "item",
                   "produto", "onde", "nome")),
    ("valor", ("valor", "valor parcela", "valor da parcela", "quantia", "montante",
               "preco", "amount", "value", "total", "valor rs", "valor r")),
)

CAMPOS = tuple(campo for campo, _ in SINONIMOS)

FORMATOS_DATA = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d/%m/%y",
    "%d-%m-%Y",
    "%Y/%m/%d",
    "%d.%m.%Y",
    "%Y-%m-%d %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y %H:%M:%S",
)


class ErroImportacao(RuntimeError):
    pass


# --------------------------------------------------------------------------
# leitura bruta
# --------------------------------------------------------------------------

def ler_tabela(caminho: Path) -> List[List[Any]]:
    caminho = Path(caminho)
    if not caminho.exists():
        raise ErroImportacao(f"arquivo não encontrado: {caminho}")
    sufixo = caminho.suffix.lower()
    if sufixo in (".xlsx", ".xlsm", ".xltx"):
        return _ler_xlsx(caminho)
    if sufixo == ".xls":
        raise ErroImportacao(
            ".xls antigo não é suportado — abra no Excel/Google Sheets e salve como .xlsx ou .csv"
        )
    return _ler_csv(caminho)


def _ler_csv(caminho: Path) -> List[List[Any]]:
    bruto = None
    for codificacao in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            bruto = caminho.read_text(encoding=codificacao)
            break
        except UnicodeDecodeError:
            continue
    if bruto is None:
        raise ErroImportacao(f"não consegui decodificar o arquivo: {caminho}")

    amostra = bruto[:8192]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t|")
        delimitador = dialeto.delimiter
    except csv.Error:
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","
    return [linha for linha in csv.reader(io.StringIO(bruto), delimiter=delimitador) if any(
        str(celula).strip() for celula in linha
    )]


def _ler_xlsx(caminho: Path) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as erro:
        raise ErroImportacao(
            "pra ler .xlsx instale o openpyxl (`pip install openpyxl`) ou exporte a "
            "planilha como CSV"
        ) from erro
    planilha = load_workbook(filename=str(caminho), read_only=True, data_only=True)
    aba = planilha.active
    linhas = []
    for linha in aba.iter_rows(values_only=True):
        if any(celula is not None and str(celula).strip() for celula in linha):
            linhas.append(list(linha))
    planilha.close()
    return linhas


# --------------------------------------------------------------------------
# detecção de colunas e conversões
# --------------------------------------------------------------------------

def detectar_colunas(cabecalho: Sequence[Any]) -> Dict[str, int]:
    """Casa cada campo conhecido com o índice da coluna correspondente."""
    normalizados = [normalizar_chave(c) for c in cabecalho]
    mapa: Dict[str, int] = {}
    usados = set()

    for campo, sinonimos in SINONIMOS:  # exato primeiro
        for indice, nome in enumerate(normalizados):
            if indice in usados or not nome:
                continue
            if nome in sinonimos:
                mapa[campo] = indice
                usados.add(indice)
                break

    for campo, sinonimos in SINONIMOS:  # depois por conteúdo
        if campo in mapa:
            continue
        for indice, nome in enumerate(normalizados):
            if indice in usados or not nome:
                continue
            if any(s in nome or nome in s for s in sinonimos):
                mapa[campo] = indice
                usados.add(indice)
                break
    return mapa


def parsear_data(bruto) -> str:
    if bruto is None or str(bruto).strip() == "":
        raise ValueError("data vazia")
    if isinstance(bruto, datetime):
        return bruto.date().isoformat()
    if isinstance(bruto, date):
        return bruto.isoformat()
    if isinstance(bruto, (int, float)) and not isinstance(bruto, bool):
        # número de série do Excel (base 1899-12-30)
        return (date(1899, 12, 30) + timedelta(days=int(bruto))).isoformat()

    texto = str(bruto).strip()
    for formato in FORMATOS_DATA:
        try:
            return datetime.strptime(texto, formato).date().isoformat()
        except ValueError:
            continue
    if texto.replace(".", "", 1).isdigit() and len(texto.split(".")[0]) <= 6:
        return (date(1899, 12, 30) + timedelta(days=int(float(texto)))).isoformat()
    raise ValueError(f"data em formato desconhecido: {bruto!r}")


def _celula(linha: Sequence[Any], mapa: Dict[str, int], campo: str):
    indice = mapa.get(campo)
    if indice is None or indice >= len(linha):
        return None
    valor = linha[indice]
    if isinstance(valor, str):
        valor = valor.strip()
    return valor if valor not in ("", None) else None


def _parcela(bruto) -> Tuple[int, int]:
    """Aceita '3/10', '10', 3 → (parcela_atual, total_parcelas)."""
    if bruto is None:
        return 1, 1
    texto = str(bruto).strip().lower().replace("x", "")
    if "/" in texto:
        partes = texto.split("/")
        try:
            return max(1, int(float(partes[0]))), max(1, int(float(partes[1])))
        except ValueError:
            return 1, 1
    try:
        total = max(1, int(float(texto)))
    except ValueError:
        return 1, 1
    return 1, total


def linha_para_lancamento(linha: Sequence[Any], mapa: Dict[str, int], criado_em: str) -> Dict[str, Any]:
    data = validar_iso(parsear_data(_celula(linha, mapa, "data")))

    bruto_valor = _celula(linha, mapa, "valor")
    if bruto_valor is None:
        bruto_valor = _celula(linha, mapa, "valorTotal")
    if bruto_valor is None:
        raise ValueError("linha sem valor")
    valor = parsear_valor(bruto_valor)
    if valor == 0:
        raise ValueError("valor zerado")

    categoria_bruta = _celula(linha, mapa, "categoria")
    tipo_bruto = _celula(linha, mapa, "tipo")
    canonica = regras.categoria_canonica_estrita(categoria_bruta)
    if tipo_bruto:
        tipo = regras.normalizar_tipo(tipo_bruto)
    elif valor < 0:
        tipo = config.TIPO_DESPESA
    elif canonica in config.CATEGORIAS_RECEITA:
        # sem coluna de tipo: a categoria denuncia ("Salário", "Outros_Receita"...)
        tipo = config.TIPO_RECEITA
    else:
        tipo = config.TIPO_DESPESA

    valor = abs(valor)
    categoria = regras.normalizar_categoria(categoria_bruta, tipo)
    forma, conta = regras.aplicar_regras(
        _celula(linha, mapa, "formaPagamento"), _celula(linha, mapa, "conta")
    )
    parcela_atual, total_parcelas = _parcela(_celula(linha, mapa, "totalParcelas"))

    bruto_total = _celula(linha, mapa, "valorTotal")
    try:
        valor_total = abs(parsear_valor(bruto_total)) if bruto_total is not None else round(
            valor * total_parcelas, 2
        )
    except ValueError:
        valor_total = round(valor * total_parcelas, 2)

    return {
        "id": uuid.uuid4().hex,
        "data": data,
        "tipo": tipo,
        "categoria": categoria,
        "valor": valor,
        "valorTotal": valor_total,
        "parcelaAtual": parcela_atual,
        "totalParcelas": total_parcelas,
        "formaPagamento": forma,
        "conta": conta,
        "descricao": str(_celula(linha, mapa, "descricao") or "").strip(),
        "criadoEm": criado_em,
        "grupoParcelamento": None,
    }


# --------------------------------------------------------------------------
# importação
# --------------------------------------------------------------------------

def importar(
    conexao,
    caminho: Path,
    *,
    simular: bool = False,
    checar_duplicatas: bool = True,
    tolerancia_dias: int = 1,
) -> Dict[str, Any]:
    linhas = ler_tabela(Path(caminho))
    if len(linhas) < 2:
        raise ErroImportacao("a planilha não tem dados além do cabeçalho")

    cabecalho, corpo = linhas[0], linhas[1:]
    mapa = detectar_colunas(cabecalho)
    if "data" not in mapa or ("valor" not in mapa and "valorTotal" not in mapa):
        raise ErroImportacao(
            "não achei as colunas obrigatórias (data e valor) no cabeçalho: "
            f"{list(cabecalho)}\ncolunas reconhecidas: {sorted(mapa)}"
        )

    criado_em = db.agora()
    novos: List[Dict[str, Any]] = []
    duplicados: List[Dict[str, Any]] = []
    erros: List[Dict[str, Any]] = []

    for numero, linha in enumerate(corpo, start=2):
        try:
            lancamento = linha_para_lancamento(linha, mapa, criado_em)
        except Exception as erro:
            erros.append({"linha": numero, "motivo": str(erro), "conteudo": list(linha)[:8]})
            continue

        if checar_duplicatas:
            existente = db.existe_duplicata(
                conexao, lancamento["data"], lancamento["valor"], tolerancia_dias=tolerancia_dias
            )
            if existente is None:
                existente = _duplicata_no_lote(novos, lancamento, tolerancia_dias)
            if existente is not None:
                duplicados.append({"linha": numero, "lancamento": lancamento, "existente": existente})
                continue

        novos.append(lancamento)

    if novos and not simular:
        db.inserir(conexao, novos)

    return {
        "arquivo": str(caminho),
        "colunasDetectadas": {campo: list(cabecalho)[indice] for campo, indice in sorted(mapa.items())},
        "colunasIgnoradas": [
            str(nome) for indice, nome in enumerate(cabecalho) if indice not in set(mapa.values())
        ],
        "linhasLidas": len(corpo),
        "importados": 0 if simular else len(novos),
        "aImportar": len(novos),
        "duplicados": duplicados,
        "erros": erros,
        "amostra": novos[:5],
        "simulado": simular,
    }


def _duplicata_no_lote(
    novos: Sequence[Dict[str, Any]], candidato: Dict[str, Any], tolerancia_dias: int
) -> Optional[Dict[str, Any]]:
    centavos = para_centavos(candidato["valor"])
    for existente in novos:
        if para_centavos(existente["valor"]) != centavos:
            continue
        if dias_entre(existente["data"], candidato["data"]) <= tolerancia_dias:
            return existente
    return None
